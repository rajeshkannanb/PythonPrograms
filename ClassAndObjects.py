# Note: There are two files combined. 
# user.py - To create classes and objects
# main.py - To use openpyxl and to read from excel.

# user.py
class User:
    def __init__(self, username, password):
        self.username = username
        self.password = password

    def set_user(self, username, password):
        self.username = username
        self.password = password

    def get_user(self):
        print(f"User Name: {self.username}, Password: {self.password}")


# main.py
import openpyxl
from user import User

inventory_workbook = openpyxl.load_workbook("D:\\Inventory.xlsx")
inventory_list = inventory_workbook["Sheet1"]

product_per_supplier = {}
product_per_inventory = {}

for product_row in range(2, inventory_list.max_row + 1):
    supplier_name = inventory_list.cell(product_row, 4).value
    print(supplier_name)

    if supplier_name in product_per_supplier:
        current_number_of_products = product_per_supplier.get(supplier_name)
        product_per_supplier[supplier_name] = current_number_of_products + 1
    else:
        product_per_supplier[supplier_name] = 1

    product_number = inventory_list.cell(product_row, 1).value
    product_inventory = inventory_list.cell(product_row, 2).value

    if product_inventory < 10:
        product_per_inventory[product_number] = product_inventory

print(product_per_inventory)
print(product_per_supplier)

user_one = User("Rajesh", "password1")
user_two = User("Kannan", "password2")

# Before changing the credentials
user_one.get_user()

# After changing the credentials
user_one.set_user("Key", "password")
user_one.get_user()

